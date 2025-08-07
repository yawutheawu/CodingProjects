'''
Maykl Yakubovsky (MC Intern Summer of 2025)
Date Created: 06/06/2025
Last Edit: 08/06/2025
Title/Function: Automation script for MC/AC Supervior daily routine task: Verifying Outstanding Priority List Keytrols
Description and Reasoning: Unfortunately, Rumba Macros are not powerful enough to allow the macro to get the keytrols from a CSV file, so a GUI automation is used instead

To-Do:


Complete:
[X] Fix R-status to only be Status not S if not in the P-List for that date, otherwise get status from scan P-List
[X] Add a togglable option that will allow automatic bypass of the select chain screen
[X] Fix the login check to check the page name instead of Scan name to allow for automatic backtracking to manager screen
[X] Blank Line Deletion
'''

import pyautogui as pg
import openpyxl as xsl
import time
import winsound
import json
import re
import pyperclip as clip    
import datetime
import socket
import os

StartTime = time.time()

# Controls how long the delay between inputs for failsafe is (0.1 by default, 0.01 seems to just barely work if the mouse is moved quickly)
#0.01 is too fast for scan to work right
pg.PAUSE = 0.05

#Attempt to automatically clear the "Select Desired Chain" screen when it appears
AutoChainSelect = True

scriptRunner = str(socket.gethostname())
print(f"Running on PC: {scriptRunner}")

pcClickLocs = {}

fileName = str(os.path.basename(__file__))
OverallFolder = str(__file__).replace(fileName,"")
os.chdir(OverallFolder)
os.chdir("Additonal")
os.chdir("Files")
with open("ScanClickLocations.json","r") as f:
    pcClickLocs = json.load(f)

if not scriptRunner in pcClickLocs.keys():
    pg.alert(text = "This PC does not have scan positions set up, and will use default values. This can lead to errors that may not be caught by the script. In case of these errors, quickly move the mouse to any corner of the screen to trigger the failsafe and force end the script.",title='Warning', button='OK')

#Globals
scanWindowName= "Worcester - Micro Focus Rumba+ Desktop"

try:
    scanTitlePos = pcClickLocs[scriptRunner]["scanTitlePos"]
except:
    scanTitlePos = pcClickLocs["Default"]["scanTitlePos"]
try:
    mainScreenOptions = pcClickLocs[scriptRunner]["mainScreenOptions"]
except:
    mainScreenOptions = pcClickLocs["Default"]["mainScreenOptions"]
try:
    headerInputPos = pcClickLocs[scriptRunner]["headerInputPos"]
except:
    headerInputPos = pcClickLocs["Default"]["headerInputPos"]
try:
    HeaderStatus = pcClickLocs[scriptRunner]["HeaderStatus"]
except:
    HeaderStatus = pcClickLocs["Default"]["HeaderStatus"]
try:
    ASNStatus = pcClickLocs[scriptRunner]["ASNStatus"]
except:
    ASNStatus = pcClickLocs["Default"]["ASNStatus"]
try:
    NumHeadersCreated = pcClickLocs[scriptRunner]["NumHeadersCreated"]
except:
    NumHeadersCreated = pcClickLocs["Default"]["NumHeadersCreated"]
try:
    ProcArea = pcClickLocs[scriptRunner]["ProcArea"]
except:
    ProcArea = pcClickLocs["Default"]["ProcArea"]
try:
    DeptNum = pcClickLocs[scriptRunner]["DeptNum"]
except:
    DeptNum = pcClickLocs["Default"]["DeptNum"]
try:
    PLdate = pcClickLocs[scriptRunner]["PLdate"]
except:
    PLdate = pcClickLocs["Default"]["PLdate"]
try:
    PLdateListStart = pcClickLocs[scriptRunner]["PLdateListStart"]
except:
    PLdateListStart = pcClickLocs["Default"]["PLdateListStart"]
try:
    PLselector = pcClickLocs[scriptRunner]["PLselector"]
except:
    PLselector = pcClickLocs["Default"]["PLselector"]
try:
    selectDesiredChain = pcClickLocs[scriptRunner]["selectDesiredChain"]
except:
    selectDesiredChain = pcClickLocs["Default"]["selectDesiredChain"]
try:
    EnterKeytrolTitle = pcClickLocs[scriptRunner]["EnterKeytrolTitle"]
except:
    EnterKeytrolTitle = pcClickLocs["Default"]["EnterKeytrolTitle"]
try:
    pListAddClick = pcClickLocs[scriptRunner]["pListAddClick"]
except:
    pListAddClick = pcClickLocs["Default"]["pListAddClick"]
try:
    pageTitle = pcClickLocs[scriptRunner]["pageTitle"]
except:
    pageTitle = pcClickLocs["Default"]["pageTitle"]
try:
    ChainInput = pcClickLocs[scriptRunner]["selectDesiredAction"]
except:
    ChainInput = pcClickLocs["Default"]["selectDesiredAction"]
try:
    PListKeytrolIn = pcClickLocs[scriptRunner]["PLISTKeytrolPosTo"]
except:
    PListKeytrolIn = pcClickLocs["Default"]["PLISTKeytrolPosTo"]
try:
    PlistFirstRowKeytrol = pcClickLocs[scriptRunner]["PListFirstRowKeytrol"]
except:
    PlistFirstRowKeytrol = pcClickLocs["Default"]["PListFirstRowKeytrol"]
try:
    PlistFirstRowStatus = pcClickLocs[scriptRunner]["PListStatus"]
except:
    PlistFirstRowStatus = pcClickLocs["Default"]["PListStatus"]
try:
    PLDateIn = pcClickLocs[scriptRunner]["PLISTPosTo"]
except:
    PLDateIn = pcClickLocs["Default"]["PLISTPosTo"]

#Highlights
# https://www.myfixguide.com/color-converter/
LStatuHighlightARGB = "fff4b084".upper()
RStatusHighlightARGB =  "ffff0000".upper() #Status Not S Highlight
whiteARGB = "ffffffff".upper()

LStatusHighlightColor = xsl.styles.colors.Color(rgb=LStatuHighlightARGB)
LStatusHighlight = xsl.styles.fills.PatternFill(start_color=LStatuHighlightARGB, end_color=LStatuHighlightARGB, fill_type='solid')
RStatusHighlightColor = xsl.styles.colors.Color(rgb=RStatusHighlightARGB)
RStatusHighlight = xsl.styles.fills.PatternFill(start_color=RStatusHighlightARGB, end_color=RStatusHighlightARGB, fill_type='solid')
whiteColor =  xsl.styles.colors.Color(rgb=whiteARGB)
whiteHighlight = xsl.styles.fills.PatternFill(start_color=whiteColor, end_color=whiteColor, fill_type='solid')

no_fill = xsl.styles.PatternFill(fill_type=None)
side = xsl.styles.Side(border_style=None)
sideLine = xsl.styles.Side(border_style="thin")
no_border = xsl.styles.borders.Border(
    left=side, 
    right=side, 
    top=side, 
    bottom=side,
)

cellBorder = xsl.styles.borders.Border(
    left=sideLine, 
    right=sideLine, 
    top=sideLine, 
    bottom=sideLine,
)


cellFont = xsl.styles.Font(name='Calibri',
                size=14,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

headerFont = xsl.styles.Font(name='Times New Roman',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

CellAlign = xsl.styles.Alignment(horizontal="center", vertical="bottom")

#Variating globals
lastPos = pg.position()
skipRows = []

# Regex
titlePattern = r"\d\d\/\d\d\/\d\d\d\d Priority List"

#Change Path to Folder with Outstanding Keytrol Excels
# Real Path to select from: Y:\AC_MC\MC Admin\PRIORITY LIST & ERROR REPORT UPDATE 2024\FY26 Outstanding Priority Keytrols
# test path: C:\Users\may02033\OneDrive - The TJX Companies, Inc\Desktop\Macro Work

folderPath =r"Y:\AC_MC\MC Admin\PRIORITY LIST & ERROR REPORT UPDATE 2024\FY26 Outstanding Priority Keytrols"

''' Set Path Relative to Macro (Makes it work with any Fiscal Year folder once dropped in)
folderPath = str(__file__).replace(str(os.path.basename(__file__)),"")
os.chdir(folderPath)
os.chdir("..")
'''

os.chdir(folderPath)
#print(pg.prompt(text=str(list(os.listdir())), title='File Selection' , default='1'))
#Set to appropriate notebook, perhaps select with above func.

excelSelections = list(os.listdir())
excelSelections.remove("Macro") #Remove Folder from Selection
messageText = "Select Correct Excel File\n"
for k,i in enumerate(excelSelections):
    messageText += f"\t{k+1}. {i}\n"

selection = -1
while (not selection < len(excelSelections)) or (not selection >= 0):
    temp = pg.prompt(text=messageText, title='Sheet Selection' , default='-1')
    try:
        selection = int(temp) - 1
        if not selection - 1 < len(excelSelections) and not selection >= 0:
            pg.alert(text = "Please Enter a Number in the list",title='Error', button='OK')
    except Exception as e:
        pg.alert(text = "Please Enter a Number",title='Error', button='OK')
        selection = -1

#excelPath = r"C:\Users\may02033\OneDrive - The TJX Companies, Inc\Desktop\Macro Work\Macro Test.xlsx"
excelPath = excelSelections[selection]


savePath = excelPath.replace(str(os.path.basename(excelPath)),"NEW_" + str(os.path.basename(excelPath)).split(".xlsx")[0] + ".xlsx")
report = xsl.load_workbook(excelPath)

#Get Correct Sheet Name:
nameList = report.sheetnames
messageText = "Select Correct Sheet Name with P-List Keytrols\n"
for k,i in enumerate(nameList):
    messageText += f"\t{k+1}. {i}\n"
selection = -1
while (not selection < len(nameList)) or (not selection >= 0):
    temp = pg.prompt(text=messageText, title='Sheet Selection' , default='-1')
    try:
        selection = int(temp) - 1
        if not selection - 1 < len(nameList) and not selection > 0:
            pg.alert(text = "Please Enter a Number in the list",title='Error', button='OK')
    except Exception as e:
        pg.alert(text = "Please Enter a Number",title='Error', button='OK')
        selection = -1

reportSheet = report[report.sheetnames[selection]]

blanksToCount = 10
rowNum = 1
blankCounter = 0
while blankCounter < blanksToCount:
    if str(reportSheet["A" + str(rowNum)].value).strip() == "" or reportSheet["A" + str(rowNum)].value == None:
        blankCounter += 1
        skipRows.append(rowNum)
    else:
        blankCounter = 0
    rowNum += 1
rowNum = rowNum-(blanksToCount+1)

#Setup Loop:
pg.getWindowsWithTitle(scanWindowName)[0].activate()
pg.getWindowsWithTitle(scanWindowName)[0].maximize()
pg.moveTo(pageTitle[0][0],pageTitle[0][1])
time.sleep(0.001)
pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
time.sleep(0.001)
pg.hotkey("ctrl","c")
while clip.paste().strip() != "Manager":
    '''
    pg.alert(text='Please Sign Into Scan! (and go to home screen)', title='User Logon Error', button='OK')
    time.sleep(10)
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    pg.dragTo(pageTitle[1][0],pageTitle[1][1])
    pg.hotkey("ctrl","c")
    '''
    pg.hotkey("f12")
    time.sleep(0.25)
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    time.sleep(0.001)
    pg.dragTo(pageTitle[1][0],pageTitle[1][1] + 1,duration=0.75)
    time.sleep(0.001)
    pg.hotkey("ctrl","c")
    time.sleep(0.25)
pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
pg.press("del",presses=2)
pg.typewrite(["2",'enter'])
pg.moveTo(pageTitle[0][0],pageTitle[0][1])
pg.dragTo(pageTitle[1][0],pageTitle[1][1], button='left',duration=0.75)
pg.hotkey('ctrl', 'c')
while clip.paste().strip() != "Processing Supervisor":
    pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
    pg.press("del",presses=2)
    pg.typewrite(["2",'enter'])
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    pg.dragTo(pageTitle[1][0],pageTitle[1][1], button='left',duration=0.75)
    pg.hotkey('ctrl', 'c')
pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
pg.press("del",presses=2)
pg.typewrite(["6",'enter'])
pg.moveTo(EnterKeytrolTitle[0][0],EnterKeytrolTitle[0][1])
pg.dragTo(EnterKeytrolTitle[1][0],EnterKeytrolTitle[1][1])
pg.hotkey('ctrl', 'c')
while clip.paste().strip() != "Enter Keytrol to Query/Edit":
    pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
    pg.press("del",presses=2)
    pg.typewrite(["6",'enter'])
    pg.moveTo(EnterKeytrolTitle[0][0],EnterKeytrolTitle[0][1])
    pg.dragTo(EnterKeytrolTitle[1][0],EnterKeytrolTitle[1][1], button='left')
    pg.hotkey('ctrl', 'c')
#Main Func Loop
CurrentPList = None
print(f"Estimated runtime: {round(((rowNum+1)*2.07)/60,2)} minutes")                                          
for i in range(1,rowNum+1):
    if len(re.findall(titlePattern,str(reportSheet["A" + str(i)].value))) > 0:
        print(f"{str(reportSheet["A" + str(i)].value)}, Row {i} of {rowNum+1}")
        CurrentPList = str(reportSheet["A" + str(i)].value).replace(" Priority List","").strip()
    elif reportSheet["A" + str(i)].value == None or str(reportSheet["A" + str(i)].value).strip() == "":
        print(f"Skipping Blank, Row {i} of {rowNum+1}")
    elif str(reportSheet["A" + str(i)].value).strip() == "Keytrol":
        print(f"Skipping Chart Headers, Row {i} of {rowNum+1}")
    else:
        hdrNum = str(reportSheet["A"+str(i)].value)
        Area = str(reportSheet["B"+str(i)].value)
        Dept = str(reportSheet["C"+str(i)].value)
        Plts = str(reportSheet["D"+str(i)].value)
        Status = str(reportSheet["E"+str(i)].value)
        Banked = str(reportSheet["F"+str(i)].value)
        print(f"Keytrol Number: {hdrNum}, Row {i} of {rowNum+1}")
        pg.click(headerInputPos[0],headerInputPos[1])
        pg.press('del', presses=6)
        clip.copy(hdrNum)
        pg.hotkey('ctrl', 'v')
        pg.press(["enter"])
        
        # Wait for input screen to clear:
        pg.press("esc")
        pg.moveTo(EnterKeytrolTitle[0][0],EnterKeytrolTitle[0][1])
        pg.dragTo(EnterKeytrolTitle[1][0],EnterKeytrolTitle[1][1])
        pg.hotkey('ctrl', 'c')
        pg.press("esc")
        dotCount = 0
        if clip.paste().strip() == 'Enter Keytrol to Query/Edit':
            print("Waiting for data.", end="")
            dotCount += 1
        while clip.paste().strip() == 'Enter Keytrol to Query/Edit':
            print(".", end="")
            dotCount+=1
            if dotCount > 10:
                print("")
            pg.press("esc")
            pg.click(headerInputPos[0],headerInputPos[1])
            clip.copy(hdrNum)
            pg.press('del', presses=6)
            pg.hotkey('ctrl', 'v')
            pg.press(["enter"])
            time.sleep(1)
            pg.press("esc")
            pg.moveTo(EnterKeytrolTitle[0][0],EnterKeytrolTitle[0][1])
            pg.dragTo(EnterKeytrolTitle[1][0],EnterKeytrolTitle[1][1])
            pg.hotkey('ctrl', 'c')
        if dotCount > 0:
            print("")
        pg.press("esc")
        #Pause on failure
        pg.moveTo(selectDesiredChain[0][0],selectDesiredChain[0][1])
        pg.dragTo(selectDesiredChain[1][0],selectDesiredChain[1][1])
        pg.hotkey('ctrl', 'c')
        pg.press("esc")
        while clip.paste().strip() == "Select Desired Chain":
            pg.moveTo(selectDesiredChain[0][0],selectDesiredChain[0][1])
            pg.dragTo(selectDesiredChain[1][0],selectDesiredChain[1][1])
            pg.hotkey('ctrl', 'c')
            if clip.paste().strip() == "Select Desired Chain":
                if AutoChainSelect:
                    pg.click(ChainInput[0],ChainInput[1])
                    pg.press(["1"])
                    pg.press(["enter"])
                else:
                    winsound.PlaySound("SystemExclamation", winsound.SND_ALIAS)
                    pg.alert(text = "Please select chain then press OK",title='Error', button='OK')
            pg.moveTo(selectDesiredChain[0][0],selectDesiredChain[0][1])
            pg.dragTo(selectDesiredChain[1][0],selectDesiredChain[1][1])
            pg.hotkey('ctrl', 'c')
        #Grab Status
        pg.press("esc")
        pg.moveTo(HeaderStatus[0][0],HeaderStatus[0][1])
        pg.dragTo(HeaderStatus[1][0],HeaderStatus[1][1])
        pg.hotkey("ctrl","c")
        KeytrolStatus = str(clip.paste()).strip()

        #Grab ASN
        pg.moveTo(ASNStatus[0][0],ASNStatus[0][1])
        pg.dragTo(ASNStatus[1][0],ASNStatus[1][1])
        pg.hotkey("ctrl","c")
        KeytrolASN = str(clip.paste()).strip()

        #Grab NumHeader
        pg.moveTo(NumHeadersCreated[0][0],NumHeadersCreated[0][1])
        pg.dragTo(NumHeadersCreated[1][0],NumHeadersCreated[1][1])
        pg.hotkey("ctrl","c")
        CreatedHeaders = str(clip.paste()).strip()

        #Grab Processing Area
        pg.moveTo(ProcArea[0][0],ProcArea[0][1])
        pg.dragTo(ProcArea[1][0],ProcArea[1][1])
        pg.hotkey("ctrl","c")
        SCAN_Area = str(clip.paste()).strip()

        #Grab Department Num
        pg.moveTo(DeptNum[0][0],DeptNum[0][1])
        pg.dragTo(DeptNum[1][0],DeptNum[1][1])
        pg.hotkey("ctrl","c")
        SCAN_Dept = str(clip.paste()).strip().replace(" ","")

        #Grab PL date
        pg.hotkey("f11")
        pg.moveTo(PLdate[0][0],PLdate[0][1])
        pg.dragTo(PLdate[1][0],PLdate[1][1])
        pg.hotkey("ctrl","c")
        SCAN_PLdate = str(clip.paste()).strip()
        pg.hotkey("f12")
        pg.hotkey("f12")
        pg.click(headerInputPos[0],headerInputPos[1])
        pg.press('del', presses=6)
        if CreatedHeaders != "" and KeytrolASN == "N":
            Banked = "Y"
        elif KeytrolASN == "Y":
            Banked = "ASN"
        else:
            Banked = ""
        if KeytrolStatus == "C":
            print(f"row {i} header {hdrNum} is complete complete")
            hdrNum = None
            Area = None

            SCAN_Area = None
            Dept = None
            Plts = None
            Status = None
            Banked = None
        elif KeytrolStatus == "L":
            Status = "L-Status"
            Banked = None
        elif KeytrolStatus == "A":
            if Area[0:3].upper() == "LUR" or Area == "CF" or Area == "GIFTA" or Area == "JWLY1":
                Status = "On List"
            elif (SCAN_PLdate == "" or SCAN_PLdate == None):
                Status == "Acknowledge"
            else:
                Status = "On List"
        elif KeytrolStatus == "R":
            Status = "Status not S"
            pg.hotkey("f12")
            pg.hotkey("f12")
            print('In Main Menu')
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["3"])
            pg.press(["enter"])
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1], duration=0.75)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != 'Bulk Supervisor':
                print("waiting on bulk screen")
                pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                pg.press("del",presses=2)
                pg.typewrite(["3"])
                pg.press(["enter"])
                time.sleep(0.1)
                pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                pg.dragTo(pageTitle[1][0],pageTitle[1][1], duration=0.75)
                pg.hotkey("ctrl","c")
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["1"])
            pg.press(["enter"])
            print('In P-List by Date')
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1], duration=0.75)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != 'Priority List Status Summary - \tT.J. Maxx':
                print("waiting on p-list screen")
                pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                pg.press("del",presses=2)
                pg.typewrite(["1"])
                pg.press(["enter"])
                time.sleep(0.1)
                pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                pg.dragTo(pageTitle[1][0],pageTitle[1][1], duration=0.75)
                pg.hotkey("ctrl","c")
            foundPList = False
            counter = 0
            mouseInterval = 15
            lookForPdate = str(CurrentPList)
            lookForPdate = datetime.datetime.strptime(lookForPdate,"%m/%d/%Y").date()
            while not foundPList:
                pg.moveTo(PLdateListStart[0][0],PLdateListStart[0][1] + (mouseInterval*counter))
                pg.dragTo(PLdateListStart[1][0],PLdateListStart[1][1] + (mouseInterval*counter))
                pg.hotkey("ctrl","c")
                if str(clip.paste()).strip() == "":
                    newDate = datetime.datetime.strptime("01/01/1825","%m/%d/%Y").date()
                else:
                    copiedDate = str(clip.paste()).strip()
                    newDate = datetime.datetime.strptime(copiedDate,"%m/%d/%y").date()
                if newDate == lookForPdate:
                    foundPList = True
                elif str(clip.paste()).strip() == "":
                    pg.hotkey("pgdn")
                    counter = 0
                else:
                    counter += 1
            pg.click(PLDateIn[0],PLDateIn[1])
            pg.press('del', presses=6)
            pg.click(x=PLselector[0],y=PLselector[1] + (mouseInterval * counter),button="left")
            pg.typewrite(["5"])
            pg.press(["enter"])
            '''
            try:
                PListKeytrolIn = pcClickLocs[scriptRunner]["PLISTPosTo"]
            except:
                PListKeytrolIn = pcClickLocs["Default"]["PLISTPosTo"]
            try:
                PlistFirstRowKeytrol = pcClickLocs[scriptRunner]["PListFirstRowKeytrol"]
            except:
                PlistFirstRowKeytrol = pcClickLocs["Default"]["PListFirstRowKeytrol"]
            try:
                PlistFirstRowStatus = pcClickLocs[scriptRunner]["PListStatus"]
            except:
                PlistFirstRowStatus = pcClickLocs["Default"]["PListStatus"]
            '''
            pg.click(PListKeytrolIn[0],PListKeytrolIn[1])
            clip.copy(hdrNum)
            pg.hotkey("ctrl","v")
            pg.press(["enter"])
            time.sleep(0.1)
            pg.moveTo(PlistFirstRowKeytrol[0][0],PlistFirstRowKeytrol[0][1])
            pg.dragTo(PlistFirstRowKeytrol[1][0],PlistFirstRowKeytrol[1][1])
            pg.hotkey("ctrl","c")
            if str(clip.paste()).strip() == str(hdrNum):
                pg.moveTo(PlistFirstRowStatus[0][0],PlistFirstRowStatus[0][1])
                pg.dragTo(PlistFirstRowStatus[1][0],PlistFirstRowStatus[1][1])
                pg.hotkey("ctrl","c")
                print("Found in P-List")
                Status = str(clip.paste()).strip()
            else:
                print("Not Found in P-List")
                Status = "Status not S"
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
            pg.hotkey("ctrl","c")
            lastClip = None
            while clip.paste().strip() != "Manager":
                if clip.paste().strip() == "" or clip.paste().strip() == lastClip or clip.paste().strip() == "Manager":
                    lastClip = clip.paste().strip()
                    time.sleep(1)
                    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                    pg.dragTo(pageTitle[1][0],pageTitle[1][1] + 1)
                    pg.hotkey("ctrl","c")
                    pg.hotkey("ctrl","c")
                    pg.hotkey("ctrl","c")
                else:
                    if clip.paste().strip() != "Manager":
                        pg.hotkey("f12")
                    time.sleep(1)
                    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                    pg.dragTo(pageTitle[1][0],pageTitle[1][1] + 1,duration=0.75)
                    pg.hotkey("ctrl","c")
                    pg.hotkey("ctrl","c")
                    pg.hotkey("ctrl","c")
                    
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["2"])
            pg.press(["enter"])
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != "Processing Supervisor":
                if clip.paste().strip() == "":
                    time.sleep(0.25)
                else:
                    pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                    pg.press("del",presses=2)
                    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                    pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
                    pg.hotkey("ctrl","c")
                    if clip.paste().strip() != "Processing Supervisor":
                        pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                        pg.typewrite(["2"])
                        pg.press(["enter"])
                    time.sleep(0.25)
                    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                    pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
                    pg.hotkey("ctrl","c")
                    time.sleep(0.25)
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["6"])
            pg.press(["enter"])
            time.sleep(0.25)
            pg.moveTo(EnterKeytrolTitle[0][0],EnterKeytrolTitle[0][1])
            pg.dragTo(EnterKeytrolTitle[1][0],EnterKeytrolTitle[1][1])
            pg.hotkey('ctrl', 'c')
            dotCount = 0
            if clip.paste().strip() != 'Enter Keytrol to Query/Edit':
                print("Waiting for data", end="")
            while clip.paste().strip() != 'Enter Keytrol to Query/Edit':
                print(".", end="")
                dotCount+=1
                time.sleep(0.25)
                pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                pg.press("del",presses=2)
                pg.typewrite(["6"])
                pg.press(["enter"])
                time.sleep(0.25)
                pg.moveTo(EnterKeytrolTitle[0][0],EnterKeytrolTitle[0][1])
                pg.dragTo(EnterKeytrolTitle[1][0],EnterKeytrolTitle[1][1])
                pg.hotkey('ctrl', 'c')
                time.sleep(0.25)
        elif  KeytrolStatus == "S":
            #pg.alert(text = f"Please add the header to the {CurrentPList} P-List (S status) and then bring page back to header #{hdrNum} (Which has been copied to the clipboard) then press OK",title='Error', button='OK')
            #clip.copy(hdrNum)
            print(f"Adding {hdrNum} to {CurrentPList} P-List")
            pg.hotkey("f12")
            pg.hotkey("f12")
            print('In Main Menu')
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["3"])
            pg.press(["enter"])
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != 'Bulk Supervisor':
                print("waiting on bulk screen")
                pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                pg.press("del",presses=2)
                pg.typewrite(["3"])
                pg.press(["enter"])
                time.sleep(0.1)
                pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
                pg.hotkey("ctrl","c")
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["1"])
            pg.press(["enter"])
            print('In P-List by Date')
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != 'Priority List Status Summary - \tT.J. Maxx':
                print("waiting on p-list screen")
                pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
                pg.press("del",presses=2)
                pg.typewrite(["1"])
                pg.press(["enter"])
                time.sleep(0.1)
                pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
                pg.hotkey("ctrl","c")
            foundPList = False
            counter = 0
            mouseInterval = 15
            lookForPdate = str(CurrentPList)
            lookForPdate = datetime.datetime.strptime(lookForPdate,"%m/%d/%Y").date()
            while not foundPList:
                pg.moveTo(PLdateListStart[0][0],PLdateListStart[0][1] + (mouseInterval*counter))
                pg.dragTo(PLdateListStart[1][0],PLdateListStart[1][1] + (mouseInterval*counter))
                pg.hotkey("ctrl","c")
                if str(clip.paste()).strip() == "":
                    newDate = datetime.datetime.strptime("01/01/1825","%m/%d/%Y").date()
                else:
                    copiedDate = str(clip.paste()).strip()
                    newDate = datetime.datetime.strptime(copiedDate,"%m/%d/%y").date()
                if newDate == lookForPdate:
                    foundPList = True
                elif str(clip.paste()).strip() == "":
                    pg.hotkey("pgdn")
                    counter = 0
                else:
                    counter += 1
            pg.click(PLDateIn[0],PLDateIn[1])
            pg.press('del', presses=6)
            pg.click(x=PLselector[0],y=PLselector[1] + (mouseInterval * counter),button="left")
            pg.typewrite(["5"])
            pg.press(["enter"])
            pg.hotkey("f6")
            pg.click(x=pListAddClick[0], y=pListAddClick[1],button="left")
            pg.press('del', presses=6)
            clip.copy(hdrNum)
            pg.hotkey("ctrl","v")
            pg.press(["enter"])
            pg.moveTo(pageTitle[0][0],pageTitle[0][1])
            pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != "Manager":
                if clip.paste().strip() == "":
                    time.sleep(0.25)
                else:
                    pg.hotkey("f12")
                    time.sleep(0.25)
                    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
                    pg.dragTo(pageTitle[1][0],pageTitle[1][1],duration=0.75)
                    pg.hotkey("ctrl","c")
                    time.sleep(0.25)
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            time.sleep(0.25)
            pg.typewrite(["2"])
            pg.press("enter")
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.click(x=mainScreenOptions[0],y=mainScreenOptions[1],button="left")
            pg.press("del",presses=2)
            pg.typewrite(["6"])
            pg.press("enter")
            Status = "On List"
            time.sleep(0.5)
            pg.click(headerInputPos[0],headerInputPos[1])
            pg.press('del', presses=6)
        if hdrNum != None and hdrNum != "None":
            reportSheet["A"+str(i)].value = int(hdrNum)
        else:
            reportSheet["A"+str(i)].value = hdrNum
        reportSheet["B"+str(i)].value = SCAN_Area
        SCAN_Dept = SCAN_Dept.strip()
        try:
            reportSheet["C"+str(i)].value = int(SCAN_Dept)
        except:
            reportSheet["C"+str(i)].value = SCAN_Dept
        if Plts != None and Plts != "None":
            try:
                reportSheet["D"+str(i)].value = int(Plts)
            except:
                reportSheet["D"+str(i)].value = str(Plts)
        else:
            reportSheet["D"+str(i)].value = None
        reportSheet["E"+str(i)].value = Status
        reportSheet["F"+str(i)].value = Banked

for i in range(rowNum,0,-1):
    hdrNum = str(reportSheet["A"+str(i)].value)
    Area = str(reportSheet["B"+str(i)].value)
    Dept = str(reportSheet["C"+str(i)].value)
    Plts = str(reportSheet["D"+str(i)].value)
    Status = str(reportSheet["E"+str(i)].value).replace(" ", "")
    Banked = str(reportSheet["F"+str(i)].value)
    print(Status)

    reportSheet["A"+str(i)].font = cellFont
    reportSheet["B"+str(i)].font = cellFont
    reportSheet["C"+str(i)].font = cellFont
    reportSheet["D"+str(i)].font = cellFont
    reportSheet["E"+str(i)].font = cellFont
    reportSheet["F"+str(i)].font = cellFont

    reportSheet["A"+str(i)].border = cellBorder
    reportSheet["B"+str(i)].border = cellBorder
    reportSheet["C"+str(i)].border = cellBorder
    reportSheet["D"+str(i)].border = cellBorder
    reportSheet["E"+str(i)].border = cellBorder
    reportSheet["F"+str(i)].border = cellBorder
    
    reportSheet["A"+str(i)].alignment = CellAlign
    reportSheet["B"+str(i)].alignment = CellAlign
    reportSheet["C"+str(i)].alignment = CellAlign
    reportSheet["D"+str(i)].alignment = CellAlign
    reportSheet["E"+str(i)].alignment = CellAlign
    reportSheet["F"+str(i)].alignment = CellAlign

    if Status == "L-Status":
        reportSheet["A"+str(i)].fill = LStatusHighlight
        reportSheet["B"+str(i)].fill = LStatusHighlight
        reportSheet["C"+str(i)].fill = LStatusHighlight
        reportSheet["D"+str(i)].fill = LStatusHighlight
        reportSheet["E"+str(i)].fill = LStatusHighlight
        reportSheet["F"+str(i)].fill = LStatusHighlight
    elif Status.lower().strip().replace(" ","") == "status not s".replace(" ",""):
        reportSheet["A"+str(i)].fill = RStatusHighlight
        reportSheet["B"+str(i)].fill = RStatusHighlight
        reportSheet["C"+str(i)].fill = RStatusHighlight
        reportSheet["D"+str(i)].fill = RStatusHighlight
        reportSheet["E"+str(i)].fill = RStatusHighlight
        reportSheet["F"+str(i)].fill = RStatusHighlight
    else:
        if reportSheet["A" + str(i)].value == None or str(reportSheet["A" + str(i)].value).strip() == "":
            reportSheet["A"+str(i)].fill = no_fill
            reportSheet["B"+str(i)].fill = no_fill
            reportSheet["C"+str(i)].fill = no_fill
            reportSheet["D"+str(i)].fill = no_fill
            reportSheet["E"+str(i)].fill = no_fill
            reportSheet["F"+str(i)].fill = no_fill

            reportSheet["A"+str(i)].border = no_border
            reportSheet["B"+str(i)].border = no_border
            reportSheet["C"+str(i)].border = no_border
            reportSheet["D"+str(i)].border = no_border
            reportSheet["E"+str(i)].border = no_border
            reportSheet["F"+str(i)].border = no_border
        elif str(reportSheet["A" + str(i)].value).strip() == "Keytrol":
            reportSheet["A"+str(i)].font = headerFont
            reportSheet["B"+str(i)].font = headerFont
            reportSheet["C"+str(i)].font = headerFont
            reportSheet["D"+str(i)].font = headerFont
            reportSheet["E"+str(i)].font = headerFont
            reportSheet["F"+str(i)].font = headerFont
        elif  "Priority".strip().lower() in str(reportSheet["A" + str(i)].value).strip().lower():
            reportSheet["A"+str(i)].font = headerFont
            reportSheet["B"+str(i)].font = headerFont
            reportSheet["C"+str(i)].font = headerFont
            reportSheet["D"+str(i)].font = headerFont
            reportSheet["E"+str(i)].font = headerFont
            reportSheet["F"+str(i)].font = headerFont
        else:
            reportSheet["A"+str(i)].fill = no_fill
            reportSheet["B"+str(i)].fill = no_fill
            reportSheet["C"+str(i)].fill = no_fill
            reportSheet["D"+str(i)].fill = no_fill
            reportSheet["E"+str(i)].fill = no_fill
            reportSheet["F"+str(i)].fill = no_fill

#Clean Blank Values:
for i in range(rowNum+1,1,-1):
    try:
        reportSheet.unmerge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    except:
        pass
    if i in skipRows:
        reportSheet["A"+str(i)].fill = no_fill
        reportSheet["B"+str(i)].fill = no_fill
        reportSheet["C"+str(i)].fill = no_fill
        reportSheet["D"+str(i)].fill = no_fill
        reportSheet["E"+str(i)].fill = no_fill
        reportSheet["F"+str(i)].fill = no_fill

        reportSheet["A"+str(i)].border = no_border
        reportSheet["B"+str(i)].border = no_border
        reportSheet["C"+str(i)].border = no_border
        reportSheet["D"+str(i)].border = no_border
        reportSheet["E"+str(i)].border = no_border
        reportSheet["F"+str(i)].border = no_border
    else:
        if reportSheet["A"+str(i)].value == "" or reportSheet["A"+str(i)].value == None:
            reportSheet.delete_rows(i)

for i in range(1,rowNum+1):
    if "Priority".strip().lower() in str(reportSheet["A" + str(i)].value).strip().lower():
        reportSheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

pg.hotkey("f12")
pg.hotkey("f12")
pg.click(mainScreenOptions[0],mainScreenOptions[1])
pg.press("del",presses=2)
report.save(savePath)
print(f"Total runtime was {round(time.time() - StartTime,2)} seconds, or {round((time.time() - StartTime)/60,2)} minutes")