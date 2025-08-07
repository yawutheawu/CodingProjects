'''
Maykl Yakubovsky (MC Intern Summer of 2025)
Date Created: 06/27/2025
Last Edit: 08/05/2025
Title/Function: Automation script for MC/AC Supervior/Admin task: Takingh P-List from SCAN and generating a formatted Excel

To-Do:

Complete:
[X] Fix the login check to check the page name instead of Scan name to allow for automatic backtracking to manager screen
'''

# Notes:
# https://openpyxl.readthedocs.io/en/stable/editing_worksheets.html
# Template Address: \Macro\Additonal\Files\Outstanding Priority Keytrols WE (TEMPLATE).xlsx


import pyautogui as pg
import openpyxl as xsl
import time
import winsound
import json
import re
import itertools
import pyperclip as clip
import datetime
import socket
import os

StartTime = time.time()

# Controls how long the delay between inputs for failsafe is (0.1 by default, 0.01 seems to just barely work if the mouse is moved quickly)
#0.01 is too fast for scan to work right
pg.PAUSE = 0.03

scriptRunner = str(socket.gethostname())
print(f"Running on PC: {scriptRunner}")

pcClickLocs = {}

def resetDir():
    fileName = str(os.path.basename(__file__))
    OverallFolder = str(__file__).replace(fileName,"")
    os.chdir(OverallFolder)
resetDir()
os.chdir("Additonal")
os.chdir("Files")
with open("ScanClickLocations.json","r") as f:
    pcClickLocs = json.load(f)

if not scriptRunner in pcClickLocs.keys():
    pg.alert(text = "This PC does not have scan positions set up, and will use default values. This can lead to errors that may not be caught by the script. In case of these errors, quickly move the mouse to any corner of the screen to trigger the failsafe and force end the script.",title='Warning', button='OK')

#Globals
scanWindowName= "Worcester - Micro Focus Rumba+ Desktop"
disallowedProcAreas = ['ADWFT', 'ADWNC', 'ADWPT', 'LURAD', 'JWS1', 'JWS2', 'LUDSJ', 'LUDSW', 'LURSJ', 'LURSW', 'EVJCO', 'EVJWL', 'LUDEW', 'LUREW', 'TJU', 'EVMAR', 'EVMCO', 'LUDEM', 'LUREM', 'MCU']

#Cell Styles:
cellFont = xsl.styles.Font(name='Calibri',
                size=14,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
CellAlign = xsl.styles.Alignment(horizontal="center", vertical="bottom")

try:
    scanTitlePos = pcClickLocs[scriptRunner]["scanTitlePos"]
except:
    scanTitlePos = pcClickLocs["Default"]["scanTitlePos"]
try:
    mainScreenOptions = pcClickLocs[scriptRunner]["mainScreenOptions"]
except:
    mainScreenOptions = pcClickLocs["Default"]["mainScreenOptions"]
try:
    headerInputPos = pcClickLocs[scriptRunner]["headerInputPos"]
except:
    headerInputPos = pcClickLocs["Default"]["headerInputPos"]
try:
    HeaderStatus = pcClickLocs[scriptRunner]["HeaderStatus"]
except:
    HeaderStatus = pcClickLocs["Default"]["HeaderStatus"]
try:
    ASNStatus = pcClickLocs[scriptRunner]["ASNStatus"]
except:
    ASNStatus = pcClickLocs["Default"]["ASNStatus"]
try:
    NumHeadersCreated = pcClickLocs[scriptRunner]["NumHeadersCreated"]
except:
    NumHeadersCreated = pcClickLocs["Default"]["NumHeadersCreated"]
try:
    ProcArea = pcClickLocs[scriptRunner]["ProcArea"]
except:
    ProcArea = pcClickLocs["Default"]["ProcArea"]
try:
    DeptNum = pcClickLocs[scriptRunner]["DeptNum"]
except:
    DeptNum = pcClickLocs["Default"]["DeptNum"]
try:
    PLdate = pcClickLocs[scriptRunner]["PLdate"]
except:
    PLdate = pcClickLocs["Default"]["PLdate"]
try:
    PLdateListStart = pcClickLocs[scriptRunner]["PLdateListStart"]
except:
    PLdateListStart = pcClickLocs["Default"]["PLdateListStart"]
try:
    PLselector = pcClickLocs[scriptRunner]["PLselector"]
except:
    PLselector = pcClickLocs["Default"]["PLselector"]
try:
    selectDesiredChain = pcClickLocs[scriptRunner]["selectDesiredChain"]
except:
    selectDesiredChain = pcClickLocs["Default"]["selectDesiredChain"]
try:
    EnterKeytrolTitle = pcClickLocs[scriptRunner]["EnterKeytrolTitle"]
except:
    EnterKeytrolTitle = pcClickLocs["Default"]["EnterKeytrolTitle"]
try:
    pListAddClick = pcClickLocs[scriptRunner]["pListAddClick"]
except:
    pListAddClick = pcClickLocs["Default"]["pListAddClick"]
try:
    pageTitle = pcClickLocs[scriptRunner]["pageTitle"]
except:
    pageTitle = pcClickLocs["Default"]["pageTitle"]
try:
    pListPosInput = pcClickLocs[scriptRunner]["PLISTPosTo"]
except:
    pListPosInput = pcClickLocs["Default"]["PLISTPosTo"]
try:
    keytrolPosInput = pcClickLocs[scriptRunner]["PLISTKeytrolPosTo"]
except:
    keytrolPosInput = pcClickLocs["Default"]["PLISTKeytrolPosTo"]
try:
    FullPLScreen = pcClickLocs[scriptRunner]["FullPLScreen"]
except:
    FullPLScreen = pcClickLocs["Default"]["FullPLScreen"]
try:
    PLFirstRowKeytrol = pcClickLocs[scriptRunner]["PListFirstRowKeytrol"]
except:
    PLFirstRowKeytrol = pcClickLocs["Default"]["PListFirstRowKeytrol"]
try:
    PLSecondRowKeytrol = pcClickLocs[scriptRunner]["PListSecondRowKeytrol"]
except:
    PLSecondRowKeytrol = pcClickLocs["Default"]["PListSecondRowKeytrol"]
try:
    PLArea = pcClickLocs[scriptRunner]["PListArea"]
except:
    PLArea = pcClickLocs["Default"]["PListArea"]
try:
    PLDept = pcClickLocs[scriptRunner]["PListDept"]
except:
    PLDept = pcClickLocs["Default"]["PListDept"]
try:
    PLPlts = pcClickLocs[scriptRunner]["PListPlts"]
except:
    PLPlts = pcClickLocs["Default"]["PListPlts"]
try:
    PLStatus = pcClickLocs[scriptRunner]["PListStatus"]
except:
    PLStatus = pcClickLocs["Default"]["PListStatus"]
try:
    PLBank = pcClickLocs[scriptRunner]["PListLURHdr"]
except:
    PLBank = pcClickLocs["Default"]["PListLURHdr"]
try:
    PLpo = pcClickLocs[scriptRunner]["PLPo"]
except:
    PLpo = pcClickLocs["Default"]["PLPo"]


mouseInterval = (PLSecondRowKeytrol[0][1] - PLFirstRowKeytrol[0][1])/2.5
secondPosInt = mouseInterval/3

# Regex
titlePattern = r"\d\d\/\d\d\/\d\d\d\d Priority List"

#Change Path to Folder with Outstanding Keytrol Excels
# Real Path to select from: Y:\AC_MC\MC Admin\PRIORITY LIST & ERROR REPORT UPDATE 2024\FY26 Outstanding Priority Keytrols
# test path: C:\Users\may02033\OneDrive - The TJX Companies, Inc\Desktop\Macro Work
folderPath =r"Y:\AC_MC\MC Admin\PRIORITY LIST & ERROR REPORT UPDATE 2024\FY26 Outstanding Priority Keytrols"

''' Set Path Relative to Macro (Makes it work with any Fiscal Year folder once dropped in)
folderPath = str(__file__).replace(str(os.path.basename(__file__)),"")
os.chdir(folderPath)
os.chdir("..")
'''

os.chdir(folderPath)

#print(pg.prompt(text=str(list(os.listdir())), title='File Selection' , default='1'))
#Set to appropriate notebook, perhaps select with above func.

noConfirm = True
while noConfirm:
    dates = pg.prompt(text='Input the 2 new P-List dates (In the format MM/DD/YYYY) split by a comma (Should be consecutive)', title='P-List dates' , default=f"{(datetime.date.today() - datetime.timedelta(1)).strftime('%m/%d/%Y')},{datetime.date.today().strftime('%m/%d/%Y')}")
    if "," in dates:
        dates = dates.split(",")
        try:
            dates[0] = datetime.datetime.strptime(str(dates[0]).strip(),"%m/%d/%Y")
            dates[1] = datetime.datetime.strptime(str(dates[1]).strip(),"%m/%d/%Y")
            dates.sort()
            confirm = pg.confirm(text=f"You have inputted {datetime.datetime.strftime(dates[0],"%m/%d/%Y")} and {datetime.datetime.strftime(dates[1],"%m/%d/%Y")}, is this correct?", title = "Confirm",  buttons=['Yes', 'No'])
            if confirm == "Yes":
                noConfirm = False
            else:
                noConfirm = True
        except:
            pg.alert(text = "Please enter 2 dates (In the format MM/DD/YYYY)",title='Warning', button='OK')
    else:
        pg.alert(text = "Please enter 2 dates split by a comma",title='Warning', button='OK')

folderPath = str(__file__).replace(str(os.path.basename(__file__)),"")
os.chdir(folderPath)
os.chdir("..")

excelPath = str(os.getcwd()) + "Outstanding Priority Keytrols WE (TEMPLATE).xlsx".replace("(TEMPLATE)",str(datetime.datetime.strftime(dates[1],"%m/%d/%Y")))
savePath = excelPath.replace(str(os.path.basename(excelPath)),"\\NEW_" + str(os.path.basename(excelPath)))
folderPath = str(__file__).replace(str(os.path.basename(__file__)),"")
os.chdir(folderPath)
os.chdir("Additonal")
os.chdir("Files")
report = xsl.load_workbook("Outstanding Priority Keytrols WE (TEMPLATE).xlsx")
reportSheet = report['Outstanding P-list WE XX.XX.XX']
reportSheet.title = f"Outstanding P-list WE {str(datetime.datetime.strftime(dates[1],"%m.%d.%y"))}" 

pg.getWindowsWithTitle(scanWindowName)[0].activate()
pg.getWindowsWithTitle(scanWindowName)[0].maximize()
pg.moveTo(pageTitle[0][0],pageTitle[0][1])
pg.dragTo(pageTitle[1][0],pageTitle[1][1])
pg.hotkey("ctrl","c")
while clip.paste().strip() != "Manager":
    pg.alert(text='Please Sign Into Scan! (and go to home screen)', title='User Logon Error', button='OK')
    time.sleep(10)
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    pg.dragTo(pageTitle[1][0],pageTitle[1][1])
    pg.hotkey("ctrl","c")
pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
pg.typewrite(["3",'enter'])
pg.moveTo(pageTitle[0][0],pageTitle[0][1])
pg.dragTo(pageTitle[1][0],pageTitle[1][1], button='left')
pg.hotkey('ctrl', 'c')
while clip.paste().strip() != "Bulk Supervisor":
    pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
    pg.typewrite(["3",'enter'])
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    pg.dragTo(pageTitle[1][0],pageTitle[1][1], button='left')
    pg.hotkey('ctrl', 'c')
pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
pg.typewrite(["1",'enter'])
pg.moveTo(pageTitle[0][0],pageTitle[0][1])
pg.dragTo(pageTitle[1][0],pageTitle[1][1])
pg.hotkey('ctrl', 'c')
while clip.paste().strip() != "Priority List Status Summary - 	T.J. Maxx":
    pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
    pg.typewrite(["1",'enter'])
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    pg.dragTo(pageTitle[1][0],pageTitle[1][1], button='left')
    pg.hotkey('ctrl', 'c')
pg.click(pListPosInput[0],pListPosInput[1])
pg.press('del', presses=6)
PLISTrows = {str(datetime.datetime.strftime(dates[0],"%m/%d/%Y")) : [], str(datetime.datetime.strftime(dates[1],"%m/%d/%Y")) : []}
for i in PLISTrows.keys():
    foundPList = False
    counter = 0
    lookForPdate = str(i)
    lookForPdate = datetime.datetime.strptime(lookForPdate,"%m/%d/%Y").date()
    while not foundPList:
        pg.moveTo(PLdateListStart[0][0],PLdateListStart[0][1] + (mouseInterval*counter))
        pg.dragTo(PLdateListStart[1][0],PLdateListStart[1][1] + (mouseInterval*counter) + secondPosInt)
        pg.hotkey("ctrl","c")
        if str(clip.paste()).strip() == "":
            newDate = datetime.datetime.strptime("01/01/1825","%m/%d/%Y").date()
        elif "\n" in str(clip.paste()).strip():
            pass
        else:
            copiedDate = str(clip.paste()).strip()
            newDate = datetime.datetime.strptime(copiedDate,"%m/%d/%y").date()
        if newDate == lookForPdate:
            foundPList = True
        elif str(clip.paste()).strip() == "":
            pg.hotkey("pgdn")
            counter = 0
        else:
            counter += 1
    pg.click(x=PLselector[0],y=PLselector[1] + (mouseInterval * counter),button="left")
    pg.typewrite(["5"])
    pg.press(["enter"])
    # In P-List for Date
    pg.click(keytrolPosInput[0],keytrolPosInput[1])
    pg.press('del', presses=6)
    oldScreen = ""
    pg.moveTo(FullPLScreen[0][0],FullPLScreen[0][1])
    pg.dragTo(FullPLScreen[1][0],FullPLScreen[1][1])
    pg.hotkey("ctrl","c")
    newScreen = clip.paste()
    pagesScanned = 0
    rows = []
    while newScreen != oldScreen:
        oldScreen = newScreen
        if pagesScanned == 0:
            print(f"On Page: {pagesScanned+1}")
            counter = 0
            pg.moveTo(PLFirstRowKeytrol[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
            pg.dragTo(PLFirstRowKeytrol[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != "":
                pg.moveTo(PLFirstRowKeytrol[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                pg.dragTo(PLFirstRowKeytrol[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                pg.hotkey("ctrl","c")
                if "\n" in clip.paste().strip():
                    pass
                else:
                    keytrolTemp = clip.paste().strip()
                    pg.moveTo(PLpo[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                    pg.dragTo(PLpo[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                    pg.hotkey("ctrl","c")
                    if clip.paste().strip()[:2] == "60":
                        pg.moveTo(PLArea[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                        pg.dragTo(PLArea[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                        pg.hotkey("ctrl","c")
                        procAreaTemp = clip.paste().strip()
                        if procAreaTemp.upper() in disallowedProcAreas:
                            pass
                        else: 
                            pg.moveTo(PLDept[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLDept[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            deptTemp = clip.paste().strip()
                            pg.moveTo(PLPlts[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLPlts[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            PltsTemp = clip.paste().strip()
                            pg.moveTo(PLStatus[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLStatus[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            StatusTemp = clip.paste().strip()
                            pg.moveTo(PLBank[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLBank[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            BankedTemp = clip.paste().strip()
                            rows.append([keytrolTemp,procAreaTemp,deptTemp,PltsTemp,StatusTemp,BankedTemp])
                counter += 1
                pg.moveTo(PLFirstRowKeytrol[0][0],PLFirstRowKeytrol[0][1] + (mouseInterval * counter))
                pg.dragTo(PLFirstRowKeytrol[1][0],PLFirstRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                pg.hotkey("ctrl","c")
        else:
            print(f"On Page: {pagesScanned+1}")
            counter = 0
            pg.moveTo(PLSecondRowKeytrol[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
            pg.dragTo(PLSecondRowKeytrol[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
            pg.hotkey("ctrl","c")
            while clip.paste().strip() != "":
                pg.moveTo(PLSecondRowKeytrol[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                pg.dragTo(PLSecondRowKeytrol[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                pg.hotkey("ctrl","c")
                if "\n" in clip.paste().strip():
                    pass
                else:
                    keytrolTemp = clip.paste().strip()
                    pg.moveTo(PLpo[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                    pg.dragTo(PLpo[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                    pg.hotkey("ctrl","c")
                    if clip.paste().strip()[:2] == "60":
                        pg.moveTo(PLArea[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                        pg.dragTo(PLArea[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                        pg.hotkey("ctrl","c")
                        procAreaTemp = clip.paste().strip()
                        if procAreaTemp.upper() in disallowedProcAreas:
                            pass
                        else:
                            pg.moveTo(PLDept[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLDept[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            deptTemp = clip.paste().strip()
                            pg.moveTo(PLPlts[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLPlts[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            PltsTemp = clip.paste().strip()
                            pg.moveTo(PLStatus[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLStatus[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            StatusTemp = clip.paste().strip()
                            pg.moveTo(PLBank[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                            pg.dragTo(PLBank[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                            pg.hotkey("ctrl","c")
                            BankedTemp = clip.paste().strip()
                            rows.append([keytrolTemp,procAreaTemp,deptTemp,PltsTemp,StatusTemp,BankedTemp])
                counter += 1
                pg.moveTo(PLSecondRowKeytrol[0][0],PLSecondRowKeytrol[0][1] + (mouseInterval * counter))
                pg.dragTo(PLSecondRowKeytrol[1][0],PLSecondRowKeytrol[1][1] + (mouseInterval * counter) + secondPosInt)
                pg.hotkey("ctrl","c")
        pagesScanned += 1
        pg.hotkey("pgdn")
        time.sleep(0.25)
        pg.moveTo(FullPLScreen[0][0],FullPLScreen[0][1])
        pg.dragTo(FullPLScreen[1][0],FullPLScreen[1][1])
        pg.hotkey("ctrl","c")
        newScreen = clip.paste()
    rows.sort()
    PLISTrows[i] = list(x for x,_ in itertools.groupby(rows))
    pg.hotkey("f12")
    pg.hotkey("f12")
    pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
    pg.typewrite(["1",'enter'])
    pg.moveTo(pageTitle[0][0],pageTitle[0][1])
    pg.dragTo(pageTitle[1][0],pageTitle[1][1])
    pg.hotkey('ctrl', 'c')
    while clip.paste().strip() != "Priority List Status Summary - 	T.J. Maxx":
        pg.click(button="left",x=mainScreenOptions[0],y=mainScreenOptions[1],clicks=1)
        pg.typewrite(["1",'enter'])
        pg.moveTo(pageTitle[0][0],pageTitle[0][1])
        pg.dragTo(pageTitle[1][0],pageTitle[1][1], button='left')
        pg.hotkey('ctrl', 'c')
    pg.click(pListPosInput[0],pListPosInput[1])
    pg.press('del', presses=6)
pg.click()
pg.hotkey("f12")
pg.hotkey("f12")
resetDir()
os.chdir("..")

reportSheet["A1"].value = f"{str(datetime.datetime.strftime(dates[0],"%m/%d/%Y"))} Priority List"
reportSheet["A6"].value = f"{str(datetime.datetime.strftime(dates[1],"%m/%d/%Y"))} Priority List"
reportSheet.unmerge_cells(start_row=6, start_column=1, end_row=6, end_column=7)
'''
# Insert a new row at the second position
ws.insert_rows(2)

# Add data to the newly inserted row
ws.cell(row=2, column=1, value="Charlie")
ws.cell(row=2, column=2, value=28)
ws.cell(row=2, column=3, value="Chicago")

# Save the workbook
wb.save("example.xlsx")
'''
startPoint = 3
nextStart = 0
for k,j in enumerate(PLISTrows[str(datetime.datetime.strftime(dates[0],"%m/%d/%Y"))]):
    print(f"On row {startPoint + k}, {str(datetime.datetime.strftime(dates[0],"%m/%d/%Y"))}")
    reportSheet.insert_rows(startPoint + k)
    #set style
    reportSheet['A' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['B' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['C' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['D' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['E' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['F' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    
    reportSheet['A' + str(startPoint + k)].font = cellFont
    reportSheet['B' + str(startPoint + k)].font = cellFont
    reportSheet['C' + str(startPoint + k)].font = cellFont
    reportSheet['D' + str(startPoint + k)].font = cellFont
    reportSheet['E' + str(startPoint + k)].font = cellFont
    reportSheet['F' + str(startPoint + k)].font = cellFont
    
    reportSheet['A' + str(startPoint + k)].alignment = CellAlign
    reportSheet['B' + str(startPoint + k)].alignment = CellAlign
    reportSheet['C' + str(startPoint + k)].alignment = CellAlign
    reportSheet['D' + str(startPoint + k)].alignment = CellAlign
    reportSheet['E' + str(startPoint + k)].alignment = CellAlign
    reportSheet['F' + str(startPoint + k)].alignment = CellAlign

    reportSheet['A' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['B' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['C' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['D' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['E' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['F' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))

    # Content:
    # [keytrolTemp,procAreaTemp,deptTemp,PltsTemp,StatusTemp,BankedTemp]
    reportSheet['A' + str(startPoint + k)].value = j[0]
    reportSheet['B' + str(startPoint + k)].value = j[1]
    reportSheet['C' + str(startPoint + k)].value = j[2]
    reportSheet['D' + str(startPoint + k)].value = j[3]
    reportSheet['E' + str(startPoint + k)].value = j[4]
    reportSheet['F' + str(startPoint + k)].value = j[5]
    nextStart = startPoint + k
#Find Next Blank post Headers
startPoint = nextStart
del nextStart
while str(reportSheet["A" + str(startPoint)].value).strip()  != f"{str(datetime.datetime.strftime(dates[1],"%m/%d/%Y"))} Priority List":
    startPoint+=1
reportSheet.merge_cells(start_row=startPoint, start_column=1, end_row=startPoint, end_column=7)
print(f"Next Blank at {startPoint}")
while reportSheet["A" + str(startPoint)].value != None:
    startPoint+=1
for k,j in enumerate(PLISTrows[str(datetime.datetime.strftime(dates[1],"%m/%d/%Y"))]):
    print(f"On row {startPoint + k}, {str(datetime.datetime.strftime(dates[0],"%m/%d/%Y"))}")
    reportSheet.insert_rows(startPoint + k)
    #set style
    reportSheet['A' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['B' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['C' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['D' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['E' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    reportSheet['F' + str(startPoint + k)].fill = xsl.styles.PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

    reportSheet['A' + str(startPoint + k)].font = cellFont
    reportSheet['B' + str(startPoint + k)].font = cellFont
    reportSheet['C' + str(startPoint + k)].font = cellFont
    reportSheet['D' + str(startPoint + k)].font = cellFont
    reportSheet['E' + str(startPoint + k)].font = cellFont
    reportSheet['F' + str(startPoint + k)].font = cellFont
    
    reportSheet['A' + str(startPoint + k)].alignment = CellAlign
    reportSheet['B' + str(startPoint + k)].alignment = CellAlign
    reportSheet['C' + str(startPoint + k)].alignment = CellAlign
    reportSheet['D' + str(startPoint + k)].alignment = CellAlign
    reportSheet['E' + str(startPoint + k)].alignment = CellAlign
    reportSheet['F' + str(startPoint + k)].alignment = CellAlign

    reportSheet['A' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['B' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['C' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['D' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['E' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))
    reportSheet['F' + str(startPoint + k)].border = xsl.styles.Border(left=xsl.styles.Side(style='thin'), right=xsl.styles.Side(style='thin'),top=xsl.styles.Side(style='thin'), bottom=xsl.styles.Side(style='thin'))

    # Content:
    # [keytrolTemp,procAreaTemp,deptTemp,PltsTemp,StatusTemp,BankedTemp]
    reportSheet['A' + str(startPoint + k)].value = j[0]
    reportSheet['B' + str(startPoint + k)].value = j[1]
    reportSheet['C' + str(startPoint + k)].value = j[2]
    reportSheet['D' + str(startPoint + k)].value = j[3]
    reportSheet['E' + str(startPoint + k)].value = j[4]
    reportSheet['F' + str(startPoint + k)].value = j[5]

#https://strftime.org/
savePath = os.getcwd() + f"\\Outstanding Priority Keytrols WE {str(datetime.datetime.strftime(dates[1],"%m.%d.%y"))}.xlsx"
report.save(savePath)
print(f"Total runtime was {round(time.time() - StartTime,2)} seconds, or {round((time.time() - StartTime)/60,2)} minutes")